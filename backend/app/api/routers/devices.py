import asyncio
import io
import logging
import re

from fastapi import APIRouter, Body, Depends, HTTPException, Query, UploadFile, File, status
from sqlalchemy import select, delete, func
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.dependencies import get_db
from app.api.schemas import DeviceCreate, DeviceOut, DeviceUpdate, DeviceWithSignals
from app.infrastructure.db.models import Device, DeviceSignal

log = logging.getLogger(__name__)
router = APIRouter(prefix="/devices", tags=["devices"])


@router.get("")
async def list_devices(
    substation_id: int | None = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(0, ge=0, le=500),
    db: AsyncSession = Depends(get_db),
):
    base_filter = []
    if substation_id is not None:
        base_filter.append(Device.substation_id == substation_id)

    # Count total
    count_q = select(func.count()).select_from(Device)
    if base_filter:
        count_q = count_q.where(*base_filter)
    total = await db.scalar(count_q)

    # Fetch items
    q = select(Device).options(selectinload(Device.signals)).order_by(Device.id).offset(skip)
    if base_filter:
        q = q.where(*base_filter)
    if limit > 0:
        q = q.limit(limit)
    result = await db.execute(q)
    return {"items": result.scalars().all(), "total": total}


@router.post("", response_model=DeviceOut, status_code=status.HTTP_201_CREATED)
async def create_device(payload: DeviceCreate, db: AsyncSession = Depends(get_db)):
    device = Device(
        substation_id=payload.substation_id,
        model_id=payload.model_id,
        name=payload.name,
        protocol=payload.protocol,
        iec104_host=payload.iec104_host,
        iec104_port=payload.iec104_port,
        iec104_common_address=payload.iec104_common_address,
        poll_interval_seconds=payload.poll_interval_seconds,
    )
    db.add(device)
    await db.flush()
    await db.refresh(device)
    return device


@router.post("/bulk-upsert")
async def bulk_upsert_devices(
    payload: list[DeviceCreate],
    db: AsyncSession = Depends(get_db),
):
    created = 0
    updated = 0
    items: list[Device] = []
    for item in payload:
        existing = await db.execute(
            select(Device).where(
                Device.iec104_host == item.iec104_host,
                Device.iec104_port == item.iec104_port,
            )
        )
        device = existing.scalars().first()
        if device:
            for field, val in item.model_dump().items():
                setattr(device, field, val)
            updated += 1
        else:
            device = Device(**item.model_dump())
            db.add(device)
            created += 1
        await db.flush()
        await db.refresh(device)
        items.append(device)

    log.info("Bulk device upsert: created=%s updated=%s", created, updated)
    return {"created": created, "updated": updated, "items": items}


@router.get("/{device_id}", response_model=DeviceWithSignals)
async def get_device(device_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Device)
        .options(selectinload(Device.signals))
        .where(Device.id == device_id)
    )
    device = result.scalar_one_or_none()
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    return device


@router.put("/{device_id}", response_model=DeviceOut)
async def update_device(
    device_id: int, payload: DeviceUpdate, db: AsyncSession = Depends(get_db)
):
    device = await db.get(Device, device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    changes = payload.model_dump(exclude_none=True)
    for field, val in changes.items():
        setattr(device, field, val)
    await db.flush()
    await db.refresh(device)
    return device


@router.delete("/{device_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_device(device_id: int, db: AsyncSession = Depends(get_db)):
    device = await db.get(Device, device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    await db.delete(device)


@router.post("/bulk-delete")
async def bulk_delete_devices(
    ids: list[int] = Body(..., embed=True),
    db: AsyncSession = Depends(get_db),
):
    if not ids:
        return {"deleted": 0}
    result = await db.execute(delete(Device).where(Device.id.in_(ids)))
    await db.flush()
    return {"deleted": result.rowcount}


@router.post("/delete-all")
async def delete_all_devices(db: AsyncSession = Depends(get_db)):
    result = await db.execute(delete(Device))
    await db.flush()
    return {"deleted": result.rowcount}


# ── Excel import ─────────────────────────────────

# IEC 104 type → value_type mapping
_ANALOG_TYPES = {"M_ME_TF_1", "M_ME_NC_1", "M_ME_NA_1", "M_ME_NB_1",
                 "M_ME_TD_1", "M_ME_TE_1", "M_ME_ND_1"}
_STATUS_TYPES = {"M_SP_TB_1", "M_SP_NA_1", "M_DP_TB_1", "M_DP_NA_1"}
_COUNTER_TYPES = {"M_IT_NA_1", "M_IT_TB_1"}
_SKIP_WORDS = {"группа", "зарезервировано", "резерв", "reserved"}


def _parse_via_excel_com(file_path: str) -> list[dict]:
    """
    Excel COM orqali faylni o'qish.
    Konfigurator-MT ning maxsus .xls formatini to'liq parse qiladi.
    """
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(file_path)
        ws = wb.Sheets(1)

        rows_count = ws.UsedRange.Rows.Count
        cols_count = min(ws.UsedRange.Columns.Count, 7)

        signals = []
        for r in range(1, rows_count + 1):
            row = []
            for c in range(1, cols_count + 1):
                val = ws.Cells(r, c).Text
                row.append(val if val else "")
            _extract_signal(row, signals)

        return signals
    finally:
        try:
            if wb:
                wb.Close(False)
            if excel:
                excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


def _parse_tsv(raw: bytes) -> list[dict]:
    """Tab-delimited text faylni o'qish (oddiy TSV)."""
    for encoding in ("utf-16", "utf-16-le", "utf-8-sig", "utf-8", "cp1251"):
        try:
            text = raw.decode(encoding)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    else:
        raise ValueError("Fayl kodlashini aniqlab bo'lmadi")

    signals = []
    for line in text.splitlines():
        parts = line.split("\t")
        if len(parts) >= 3:
            _extract_signal(parts, signals)
    return signals


def _parse_xls(raw: bytes) -> list[dict]:
    """Eski .xls faylni o'qish (xlrd)."""
    import xlrd
    wb = xlrd.open_workbook(file_contents=raw)
    ws = wb.sheet_by_index(0)
    signals = []
    for r in range(ws.nrows):
        row = [ws.cell_value(r, c) for c in range(min(ws.ncols, 7))]
        _extract_signal(row, signals)
    return signals


def _parse_xlsx(raw: bytes) -> list[dict]:
    """Yangi .xlsx faylni o'qish (openpyxl)."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    signals = []
    for row in ws.iter_rows(max_col=7, values_only=True):
        _extract_signal(list(row or []), signals)
    return signals


def _extract_signal(row: list, out: list[dict]):
    """Bitta qatordan signal ma'lumotini olish."""
    if len(row) < 3:
        return
    ioa_raw = row[0]
    name = str(row[1] or "").strip().strip('"')
    asdu_type = str(row[2] or "").strip().strip("<> ").strip()

    # IOA ni raqamga aylantirish
    try:
        ioa = int(float(ioa_raw))
    except (ValueError, TypeError):
        return

    # Guruh sarlavhalarini va zaxiralarni o'tkazish
    name_l = name.lower()
    if not name or any(w in name_l for w in _SKIP_WORDS):
        return
    if name.startswith("[- ") or name_l.startswith("[-"):
        return

    # Tur aniqlash
    if asdu_type in _ANALOG_TYPES:
        value_type = "float"
    elif asdu_type in _STATUS_TYPES:
        value_type = "status"
    elif asdu_type in _COUNTER_TYPES:
        value_type = "counter"
    else:
        # Noma'lum tur — o'tkazib yuboramiz
        return

    # Birlik aniqlash (nom ichidan)
    unit = ""
    if "," in name:
        parts = name.rsplit(",", 1)
        if len(parts) == 2 and len(parts[1].strip()) <= 8:
            unit = parts[1].strip()

    # signal_name: lotin harflarga aylantirish
    code = name
    # Maxsus belgilarni tozalash
    for ch in ["^", "|", "(", ")", " ", "/", ":", ".", ",", "°", "ф"]:
        code = code.replace(ch, "_")
    code = code.strip("_").lower()
    # Ikki pastki chiziq ketma-ket bo'lmasin
    while "__" in code:
        code = code.replace("__", "_")

    prefix = {"float": "ai", "status": "st", "counter": "cnt"}[value_type]
    code = re.sub(r"[^a-zA-Z0-9_]+", "_", f"{prefix}_{ioa}").strip("_").lower()

    out.append({
        "register_code": ioa,
        "signal_name": code[:64],
        "signal_title": name[:128],
        "unit": unit[:16],
        "value_type": value_type,
    })


@router.post("/{device_id}/signals/import-excel")
async def import_signals_from_excel(
    device_id: int,
    file: UploadFile = File(...),
    replace_existing: bool = Query(True),
    db: AsyncSession = Depends(get_db),
):
    """
    Конфигуратор-МТ дан экспорт qilingan .xls/.xlsx файлни юклаб,
    сигналларни автоматик импорт қилади.
    Мавжуд IOA лар ўтказиб юборилади.
    """
    device = await db.get(Device, device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Qurilma topilmadi")

    import os, tempfile

    raw = await file.read()
    fname = (file.filename or "unknown.xls").lower()

    # Faylni vaqtinchalik diskka yozish (Excel COM uchun)
    suffix = ".xlsx" if fname.endswith(".xlsx") else ".xls"
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=suffix, prefix="iec104_import_")
    try:
        os.write(tmp_fd, raw)
        os.close(tmp_fd)
        abs_path = os.path.abspath(tmp_path)

        signals: list[dict] = []
        try:
            if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
                signals = _parse_tsv(raw)
            elif raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
                signals = _parse_xls(raw)
            elif fname.endswith(".xlsx"):
                signals = _parse_xlsx(raw)
        except Exception as e_fast:
            log.warning("Tez parser xato: %s", e_fast)

        if not signals:
            try:
                signals = await asyncio.get_event_loop().run_in_executor(
                    None, _parse_via_excel_com, abs_path
                )
            except Exception as e_com:
                log.exception("Excel COM ham xato berdi")
                raise HTTPException(
                    status_code=400,
                    detail=f"Faylni o'qib bo'lmadi: {e_com}",
                )
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    if not signals:
        raise HTTPException(status_code=400, detail="Faylda signal topilmadi")

    # Mavjud IOAlarni olish
    existing = await db.execute(
        select(DeviceSignal).where(DeviceSignal.device_id == device_id)
    )
    existing_by_ioa = {sig.register_code: sig for sig in existing.scalars().all()}

    applied = 0
    updated = 0
    skipped = 0
    seen_ioas: set[int] = set()
    for sig in signals:
        if sig["register_code"] in seen_ioas:
            skipped += 1
            continue
        seen_ioas.add(sig["register_code"])
        current = existing_by_ioa.get(sig["register_code"])
        if current and not replace_existing:
            skipped += 1
            continue
        if current:
            current.signal_name = sig["signal_name"]
            current.signal_title = sig["signal_title"]
            current.unit = sig["unit"]
            current.value_type = sig["value_type"]
            updated += 1
            continue
        db.add(DeviceSignal(
            device_id=device_id,
            register_code=sig["register_code"],
            signal_name=sig["signal_name"],
            signal_title=sig["signal_title"],
            unit=sig["unit"],
            value_type=sig["value_type"],
        ))
        applied += 1

    await db.flush()
    log.info(
        "Excel import device_id=%s file=%s parsed=%s applied=%s updated=%s skipped=%s replace_existing=%s",
        device_id, file.filename, len(signals), applied, updated, skipped, replace_existing,
    )
    return {
        "total_in_file": len(signals),
        "applied": applied,
        "updated": updated,
        "skipped": skipped,
        "device_id": device_id,
        "device_name": device.name,
    }
